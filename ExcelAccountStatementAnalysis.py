import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import string
debit_row = 0
debit_column = 0
credit_row = 0
credit_column = 0
searchString1 = ['debit', 'Debit', 'DEBIT', 'Withdrawals']
searchString2 = ['credit', 'Credit', 'CREDIT', 'Lodgments']

Listfile = input('Please enter the full name of the Excel file to open: ' )+'.xlsx'
# Define variable to load the wookbook
workbook = openpyxl.load_workbook(Listfile)

# Define variable to read the active sheet:
worksheet = workbook.active

# how to determine how many rows and column are in the excel
max_column = worksheet.max_column
max_row = worksheet.max_row

#Class to locate cells in Excel sheet
class cell_location:
    def __init__ (self, searchString):
        self.searchString = searchString

    #Iterate the loop to get a cell's row
    def get_row(self):
        for row in range(1, max_row + 1):
            for column in range(1, max_column + 1):
                if worksheet.cell(row, column).value == self.searchString:
                    return row

    #Iterate the loop to get the column for debit and creadit
    def get_row_and_column(self):
        returnlist = []
        for row in range(1, max_row + 1):
            for column in range(1, max_column + 1):
                if worksheet.cell(row, column).value in self.searchString:
                    returnlist.append(row)
                    returnlist.append(column)
                    return returnlist
            
            
#Convert column to excel format
def excelformat(row, column):
    alp = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    cell = alp[column-1]+str(row+1)
    return cell


#class to Loop and collect values
class ite_column:
    def __init__(self, fromA):
        self.fromA = fromA
        self.toB = fromA    

    #Loop and create a list
    def createlist(self):

        listnumber = []
        for i in range(debit_row, max_row):
            for col in worksheet.iter_cols(self.fromA,self.toB):
                if str(col[i].value) != 'None':
                    listnumber.append(float(col[i].value))
        return listnumber

#To find the Average
def average(num, row_range):
    av_num = num/row_range
    return round(av_num, 2)


debit_row = cell_location(searchString1)
debit_column = cell_location(searchString1)
debit_row = debit_row.get_row_and_column()[0]
debit_column = debit_column.get_row_and_column()[1]

credit_row = cell_location(searchString2)
credit_column = cell_location(searchString2)
credit_row = credit_row.get_row_and_column()[0]
credit_column = credit_column.get_row_and_column()[1]

debit_list = ite_column(debit_column)
credit_list = ite_column(credit_column)
debit_list = debit_list.createlist()
credit_list = credit_list.createlist()

Total_Debit = sum(debit_list)
Total_Credit = sum(credit_list)
Total_Debit = round(Total_Debit, 2)
Total_Credit = round(Total_Credit, 2)
debit_range = len(debit_list)
credit_range = len(credit_list)
max_debit = max(debit_list)
max_credit = max(credit_list)

max_debit_row = cell_location(max_debit)
max_credit_row = cell_location(max_credit)
max_debit_row = max_debit_row.get_row()
max_credit_row = max_credit_row.get_row()

average_total_debit = average(Total_Debit, debit_range)
average_total_credit = average(Total_Credit, credit_range)

print("Total Debit: " + str(Total_Debit))
print("Total Credit: "+ str(Total_Credit))

print("Average Debit: " + str(average_total_debit))
print("Average Credit: "+ str(average_total_credit))

print("Max Debit: " + str(max_debit) + "  Row: " + str(max_debit_row))
print("Max Credit: " +str(max_credit) + "   Row: " +str(max_credit_row) )



