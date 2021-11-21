import openpyxl
from openpyxl import load_workbook
debit_row = 0
debit_column = 0
credit_row = 0
credit_column = 0
searchString1 = ['debit', 'Debit', 'DEBIT', 'Withdrawals']
searchString2 = ['credit', 'Credit', 'CREDIT', 'Lodgments']

# Define variable to load the wookbook
workbook = openpyxl.load_workbook("MyList.xlsx")

# how to determine how many rows and column are in the excel
max_column = workbook.active.max_column
max_row = workbook.active.max_row

print ('Max_column::::::::: ', max_column)
print ('Max_row::::::::: ', max_row)

# Define variable to read the active sheet:
worksheet = workbook.active


#Iterate the loop to get the column for debit and creadit

for i in range(1, worksheet.max_row + 1):
    for j in range(1, worksheet.max_column + 1):
        if worksheet.cell(i,j).value in searchString1:
            debit_row = i
            debit_column = j
            print('Debit in Column: ',debit_column, 'Row:', debit_row)
           
        if worksheet.cell(i,j).value in searchString2:
            credit_row = i
            credit_column = j
            print('Credit in Column: ',credit_column, 'Row:',credit_row)
            
            
#Convert column to excel format
alp = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
debit_cell_start = alp[debit_column-1]+str(debit_row+1)
credit_cell_start = alp[credit_column-1]+str(credit_row+1)
debit_cell_end = alp[debit_column-1]+str(worksheet.max_row)
credit_cell_end = alp[credit_column-1]+str(worksheet.max_row)


print ('Debit cell starts here: ', debit_cell_start)
print ('Credit cell starts here: ',credit_cell_start)
print ('Debit cell ends here: ',debit_cell_end)
print ('Credit cell ends here: ',credit_cell_end)

